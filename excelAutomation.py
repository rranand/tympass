'''
Run this command
python3 excelAutomation.py sampleData.xlsx otherSampleData.xlsx
'''



import openpyxl
import pandas as pd
import sys
from datetime import date

#path = input("Enter Path: ")
path = sys.argv[1]
otherFile = sys.argv[2]

try:
    wbObj = openpyxl.load_workbook(path)
except IOError:
    print('First File Not Found')
    exit(0)

otherFileFound = True
try:
    wbObj_Other = pd.read_excel(otherFile, engine='openpyxl')
except IOError:
    otherFileFound = False
    print('Second File Not Found')

sheet = wbObj.active

arr = []
for i in sheet.iter_rows(min_row=2, values_only=True):
    txt = i[0].startswith("CTASK")

    if not txt:
        arr.append(i)

sheet.delete_rows(2, sheet.max_row)
arr.sort(key = lambda row: row[1])

for i in arr:
    sheet.append(i)

col_mx_value = sheet.max_column+1
sheet.insert_cols(col_mx_value)
sheet.cell(1, sheet.max_column+1).value = "Aging Days"
today = date.today()

for i in range(2, sheet.max_row+1):
    cellDateTime = sheet.cell(i, 2).value
    shellDate = date(cellDateTime.year, cellDateTime.month, cellDateTime.day)
    diff = (today - shellDate).days
    weekDay = today.weekday()

    if weekDay == 0:
        if diff <= 4:
            sheet.cell(i, col_mx_value).value = "0-2 days"
        elif diff <= 7:
            sheet.cell(i, col_mx_value).value = "2-4 days"
        elif diff <= 13:
            sheet.cell(i, col_mx_value).value = "5-10 days"
        elif diff <= 23:
            sheet.cell(i, col_mx_value).value = "11-20 days"
        elif diff <= 33:
            sheet.cell(i, col_mx_value).value = "21-30 days"
        else:
            sheet.cell(i, col_mx_value).value = "More than 30 days"
    else:
        if diff <= 2:
            sheet.cell(i, col_mx_value).value = "0-2 days"
        elif diff <= 5:
            sheet.cell(i, col_mx_value).value = "2-4 days"
        elif diff <= 11:
            sheet.cell(i, col_mx_value).value = "5-10 days"
        elif diff <= 21:
            sheet.cell(i, col_mx_value).value = "11-20 days"
        elif diff <= 31:
            sheet.cell(i, col_mx_value).value = "21-30 days"
        else:
            sheet.cell(i, col_mx_value).value = "More than 30 days"


arr.clear()

columnNames = []

for i in sheet.iter_rows(values_only=True):
   arr.append(i)

for i in arr[0]:
    columnNames.append(i)

df = pd.DataFrame(arr)
df.columns = columnNames

df.drop(axis=1, index= 0, inplace=True)

pvTable = df.pivot_table(index='Assignment Group', columns='Aging Days', values='Assigned To', aggfunc='count', margins=True, margins_name='Grand Total')

df['copy_Assigned To'] = df.loc[:, 'Assigned To']
pvTableAssociate = df.pivot_table(index='Assigned To', columns='Task type', values='copy_Assigned To', aggfunc='count', margins=True, margins_name='Grand Total')
df.drop(columns=['copy_Assigned To'], inplace=True)

fileName = "".join(path.split(".")[:-1])

if otherFileFound:
    wbObj_Other['copy_Assigned To'] = wbObj_Other.loc[:, 'Assigned To']
    pvTableNotUpdateAssociate = wbObj_Other.pivot_table(index='Assigned To', columns='Status', values='copy_Assigned To', aggfunc='count', margins=True, margins_name='Grand Total')
    wbObj_Other.drop(columns=['copy_Assigned To'], inplace=True) 


with pd.ExcelWriter(fileName+"_output.xlsx") as writer:
    df.to_excel(writer, sheet_name="Raw Sheet")
    pvTable.to_excel(writer, sheet_name="Pivot", startrow=2)
    pvTableAssociate.to_excel(writer, sheet_name="Associate", startrow=2)

    if otherFileFound:
        wbObj_Other.to_excel(writer, sheet_name="Not Updated Raw")
        pvTableNotUpdateAssociate.to_excel(writer, sheet_name="Not Update Associate", startrow=2)
        

print("File Generated Successfully: "+fileName+"_output.xlsx")
